#!/usr/bin/env python3
"""
Script to generate products_template.xlsx for product data management.

This script creates a professional Excel template with:
- Products sheet with example data (16 rows covering Instagram, YouTube, TikTok, Xbox, ChatGPT)
- Instructions sheet with step-by-step guide for users
- Column Guide sheet explaining each field in detail

Run this script to regenerate the template if needed:
    python3 create_template.py

The generated file is ready for users to fill in and export as CSV for import via Discord.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_products_template():
    """Create the products template Excel file."""
    wb = Workbook()
    
    # Sheet 1: Products
    ws_products = wb.active
    ws_products.title = "Products"
    
    # Headers
    headers = [
        "Main_Category", "Sub_Category", "Service_Name", "Variant_Name", 
        "Price_USD", "Start_Time", "Duration", "Refill_Period", "Additional_Info"
    ]
    ws_products.append(headers)
    
    # Format headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_products.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Example data rows
    example_data = [
        # Instagram
        ["Instagram", "Followers", "Instagram Services", "500 Followers", 5.99, "10-25min", "100min", "30 day", "High quality, real looking accounts"],
        ["Instagram", "Followers", "Instagram Services", "1000 Followers", 10.99, "15-30min", "120min", "30 day", "Premium quality with engagement"],
        ["Instagram", "Followers", "Instagram Services", "5000 Followers", 45.00, "30-60min", "240min", "60 day", "Takes long to start - best quality"],
        ["Instagram", "Likes", "Instagram Services", "500 Likes", 3.99, "5-15min", "45min", "30 day", "Instant boost for your posts"],
        ["Instagram", "Likes", "Instagram Services", "2000 Likes", 12.99, "10-20min", "80min", "30 day", "Natural delivery pattern"],
        ["Instagram", "Comments", "Instagram Services", "50 Custom Comments", 25.00, "24-48hr", "N/A", "No refill", "Provide comment list after purchase"],
        
        # YouTube
        ["YouTube", "Subscribers", "YouTube Growth", "500 Subscribers", 29.99, "1-3hr", "72hr", "30 day", "Active accounts with profile pictures"],
        ["YouTube", "Subscribers", "YouTube Growth", "1000 Subscribers", 54.99, "2-6hr", "96hr", "60 day", "Premium service - monetization ready"],
        ["YouTube", "Watch Time", "YouTube Growth", "1000 Hours Watch Time", 149.99, "3-7 days", "N/A", "No refill", "Real views from various sources"],
        ["YouTube", "Likes", "YouTube Growth", "500 Likes", 8.99, "1-2hr", "48hr", "30 day", "Quick delivery, natural retention"],
        
        # TikTok
        ["TikTok", "Followers", "TikTok Services", "1000 Followers", 8.99, "10-30min", "120min", "30 day", "Real looking accounts"],
        ["TikTok", "Likes", "TikTok Services", "1000 Likes", 4.99, "5-15min", "60min", "30 day", "Fast delivery for viral boost"],
        ["TikTok", "Views", "TikTok Services", "10000 Views", 6.99, "1-6hr", "24hr", "No refill", "Helps push content to FYP"],
        
        # Xbox
        ["Xbox", "Game Pass", "Xbox Services", "Xbox Game Pass Ultimate - 1 Month", 9.99, "Instant", "N/A", "No refill", "Digital code delivery via DM"],
        ["Xbox", "Game Pass", "Xbox Services", "Xbox Game Pass Ultimate - 3 Months", 27.99, "Instant", "N/A", "No refill", "Best value - digital code"],
        
        # ChatGPT
        ["ChatGPT", "Subscription", "AI Services", "ChatGPT Plus - 1 Month", 18.99, "1-24hr", "N/A", "No refill", "Full GPT-4 access + plugins"],
    ]
    
    for row_data in example_data:
        ws_products.append(row_data)
    
    # Auto-fit columns
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws_products[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_products.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze header row
    ws_products.freeze_panes = "A2"
    
    # Sheet 2: Instructions
    ws_instructions = wb.create_sheet("Instructions")
    ws_instructions.column_dimensions['A'].width = 80
    
    instructions_title = "HOW TO USE THIS TEMPLATE"
    ws_instructions.append([instructions_title])
    ws_instructions['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws_instructions.append([])
    
    instructions = [
        "STEP 1: Review the Example Data",
        "   • Go to the 'Products' sheet and review the example rows",
        "   • Notice the format and structure of each column",
        "   • Check the 'Column Guide' sheet for detailed explanations",
        "",
        "STEP 2: Delete Example Rows",
        "   • Select all example rows (rows 2-17 on Products sheet)",
        "   • Right-click and choose 'Delete'",
        "   • Keep the header row (row 1) intact",
        "",
        "STEP 3: Add Your Product Data",
        "   • Start entering your products from row 2",
        "   • Follow the same column format as the examples",
        "   • Be consistent with your Main_Category and Sub_Category naming",
        "",
        "STEP 4: Format Guidelines",
        "   • Price_USD: Enter as decimal number (e.g., 20 for $20.00, 5.99 for $5.99)",
        "   • Start_Time: Examples - '10-25min', '1-3hr', '24-48hr', 'Instant'",
        "   • Duration: Examples - '100min', '72hr', 'N/A' for instant delivery",
        "   • Refill_Period: Examples - '30 day', '60 day', 'No refill'",
        "   • Additional_Info: Any special notes or instructions",
        "",
        "STEP 5: Save as Excel",
        "   • File → Save to keep your work as .xlsx",
        "   • This preserves formatting and all sheets",
        "",
        "STEP 6: Convert to CSV for Import",
        "   • When ready to import to Discord bot:",
        "   • File → Save As",
        "   • Choose 'CSV (Comma delimited) (*.csv)' format",
        "   • Only the active sheet will be saved (make sure Products sheet is active)",
        "   • Click 'Yes' when Excel warns about features",
        "",
        "STEP 7: Upload to Discord",
        "   • Use the /import_products command in Discord",
        "   • Attach your CSV file",
        "   • The bot will process and add all products",
        "",
        "EXAMPLE ROW:",
        "Main_Category: Instagram",
        "Sub_Category: Followers",
        "Service_Name: Instagram Services",
        "Variant_Name: 1000 Followers",
        "Price_USD: 10.99",
        "Start_Time: 15-30min",
        "Duration: 120min",
        "Refill_Period: 30 day",
        "Additional_Info: Premium quality with engagement",
        "",
        "TIPS:",
        "   • Use consistent naming for categories to keep products organized",
        "   • Variant_Name is what customers will see - make it clear and descriptive",
        "   • Price will be converted to cents internally (10.99 becomes 1099 cents)",
        "   • Additional_Info can include delivery notes, quality info, or requirements",
    ]
    
    for instruction in instructions:
        ws_instructions.append([instruction])
    
    # Sheet 3: Column Guide
    ws_guide = wb.create_sheet("Column Guide")
    ws_guide.column_dimensions['A'].width = 20
    ws_guide.column_dimensions['B'].width = 60
    
    guide_title = "COLUMN REFERENCE GUIDE"
    ws_guide.append([guide_title])
    ws_guide['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws_guide.merge_cells('A1:B1')
    ws_guide.append([])
    
    # Column guide headers
    ws_guide.append(["Column Name", "Description & Examples"])
    ws_guide['A3'].font = Font(bold=True, size=11)
    ws_guide['B3'].font = Font(bold=True, size=11)
    ws_guide.append([])
    
    column_guide = [
        ["Main_Category", "The primary platform or service type. This is the top-level grouping.\nExamples: Instagram, YouTube, TikTok, Xbox, ChatGPT, Netflix, Spotify, Discord"],
        ["", ""],
        ["Sub_Category", "The specific service type within the main category.\nExamples:\n  • For Instagram: Followers, Likes, Comments, Views, Story Views\n  • For YouTube: Subscribers, Watch Time, Likes, Views\n  • For TikTok: Followers, Likes, Views, Shares\n  • For Xbox: Game Pass, Live Gold\n  • For ChatGPT: Subscription, API Credits"],
        ["", ""],
        ["Service_Name", "Internal grouping name for related products. Used for organization.\nExamples:\n  • 'Instagram Services' for all Instagram products\n  • 'YouTube Growth' for YouTube packages\n  • 'TikTok Services' for TikTok products\n  • 'Xbox Services' for gaming subscriptions"],
        ["", ""],
        ["Variant_Name", "The exact product name that customers will see. Be specific and descriptive.\nExamples:\n  • '500 Followers' or '1000 Followers'\n  • 'Xbox Game Pass Ultimate - 1 Month'\n  • 'ChatGPT Plus - 1 Month'\n  • '10000 TikTok Views'"],
        ["", ""],
        ["Price_USD", "Product price in US Dollars as a decimal number.\nThe system converts this to cents internally (multiply by 100).\nExamples:\n  • 5.99 (becomes 599 cents)\n  • 10 or 10.00 (becomes 1000 cents)\n  • 149.99 (becomes 14999 cents)"],
        ["", ""],
        ["Start_Time", "When the service delivery begins after purchase.\nExamples:\n  • '10-25min' - starts within 10-25 minutes\n  • '1-3hr' - starts within 1-3 hours\n  • '24-48hr' - starts within 1-2 days\n  • 'Instant' - delivered immediately"],
        ["", ""],
        ["Duration", "How long the delivery process takes to complete.\nExamples:\n  • '100min' - completed in 100 minutes\n  • '72hr' - takes up to 72 hours\n  • 'N/A' - instant delivery or not applicable"],
        ["", ""],
        ["Refill_Period", "Guarantee/warranty period for the service.\nExamples:\n  • '30 day' - 30-day refill guarantee\n  • '60 day' - 60-day refill guarantee\n  • 'No refill' - one-time service, no guarantee"],
        ["", ""],
        ["Additional_Info", "Extra details, requirements, or special notes about the product.\nExamples:\n  • 'High quality, real looking accounts'\n  • 'Takes long to start - best quality'\n  • 'Digital code delivery via DM'\n  • 'Provide comment list after purchase'\n  • 'Real views from various sources'"],
    ]
    
    for row_data in column_guide:
        ws_guide.append(row_data)
        if row_data[0]:  # If column name is present
            ws_guide.cell(row=ws_guide.max_row, column=1).font = Font(bold=True, color="2E75B6")
    
    # Wrap text for better readability
    for row in ws_guide.iter_rows(min_row=5, max_row=ws_guide.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save the workbook
    wb.save("products_template.xlsx")
    print("✓ products_template.xlsx created successfully!")
    print("  • Sheet 1: Products (with 16 example rows)")
    print("  • Sheet 2: Instructions (step-by-step guide)")
    print("  • Sheet 3: Column Guide (detailed field explanations)")

if __name__ == "__main__":
    create_products_template()
